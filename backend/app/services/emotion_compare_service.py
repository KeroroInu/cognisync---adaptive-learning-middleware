"""
Emotion Compare Service - 情感对比实验数据集标准化与接口骨架
"""
import csv
import io
import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

from app.schemas.admin.emotion_compare import (
    EmotionCompareAnalyzeResponse,
    EmotionCompareBaselineRow,
    EmotionCompareComparisonRow,
    EmotionCompareDatasetInfo,
    EmotionCompareExportArtifacts,
    EmotionCompareProfileSnapshot,
    EmotionCompareSummaryMetrics,
    EmotionCompareSystemRow,
    EmotionCompareSingleMetrics,
    EmotionCompareMultiMetrics,
    EmotionComparePrediction,
    EmotionCompareDelta,
    EmotionCompareContextUsed,
    EmotionComparePerLabelMetric,
)
from app.schemas.admin.emotion_compare_v2 import (
    EmotionCompareAnalyzeResponseV2,
    EmotionCompareAgreementMetrics,
    EmotionCompareAgreementMetricSet,
    EmotionCompareConfusionAnalysis,
    EmotionCompareIntensityMetrics,
    EmotionCompareIntensityMetricSet,
    EmotionComparePerClassMetricItem,
)
from app.schemas.profile import UserProfile
from app.services.emotion_dataset_service import normalize_label_mapping
from app.services.emotion_compare_mapping import build_default_label_mapping
from app.services.llm_provider import BaseProvider
from app.services.text_analyzer import TextAnalyzer


LabelMode = Literal["single_label", "multi_binary"]
logger = logging.getLogger(__name__)


@dataclass(slots=True)
class NormalizedCompareRow:
    row_index: int
    sample_id: str | None
    text: str
    ground_truth_labels: list[str]


@dataclass(slots=True)
class NormalizedDataset:
    dataset_name: str
    source_format: str
    task_type: Literal["single_label", "multi_label"]
    dataset_template: str | None
    sample_id_column: str | None
    text_column: str
    expected_label_column: str | None
    expected_label_columns: list[str]
    positive_label_value: str | None
    rows_processed: int
    rows_skipped: int
    label_count: int
    labels: list[str]
    rows: list[NormalizedCompareRow]


def _default_profile() -> UserProfile:
    return UserProfile(cognition=50, affect=50, behavior=50, lastUpdate=None)


def _apply_delta(profile: UserProfile, cognition: int, affect: int, behavior: int) -> UserProfile:
    return UserProfile(
        cognition=max(0, min(100, profile.cognition + cognition)),
        affect=max(0, min(100, profile.affect + affect)),
        behavior=max(0, min(100, profile.behavior + behavior)),
        lastUpdate=profile.lastUpdate,
    )


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_label(value: Any) -> str | None:
    normalized = _normalize_cell(value)
    return normalized if normalized else None


def _load_csv_records(raw_bytes: bytes) -> tuple[list[dict[str, str]], list[str]]:
    try:
        content = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV must be UTF-8 encoded") from exc

    reader = csv.DictReader(io.StringIO(content))
    fieldnames = list(reader.fieldnames or [])
    if not fieldnames:
        raise ValueError("CSV is missing header row")
    records = [{key: _normalize_cell(value) for key, value in row.items()} for row in reader]
    return records, fieldnames


def _load_xlsx_records(raw_bytes: bytes) -> tuple[list[dict[str, str]], list[str]]:
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Failed to read XLSX file") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ValueError("XLSX file is empty")

    headers = [_normalize_cell(cell) for cell in rows[0]]
    if not any(headers):
        raise ValueError("XLSX is missing header row")

    fieldnames = [header for header in headers if header]
    if not fieldnames:
        raise ValueError("XLSX is missing valid header columns")

    records: list[dict[str, str]] = []
    for raw_row in rows[1:]:
        record: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = _normalize_cell(raw_row[index] if index < len(raw_row) else "")
        records.append(record)
    return records, fieldnames


def _load_json_records(raw_bytes: bytes) -> tuple[list[dict[str, str]], list[str]]:
    try:
        content = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("JSON/TXT dataset must be UTF-8 encoded") from exc

    try:
        payload = json.loads(content)
    except json.JSONDecodeError as exc:
        raise ValueError("TXT/JSON dataset must contain a JSON array of objects") from exc

    if not isinstance(payload, list):
        raise ValueError("TXT/JSON dataset must contain a JSON array")

    records: list[dict[str, str]] = []
    fieldnames: list[str] = []
    seen_fields: set[str] = set()
    for item in payload:
        if not isinstance(item, dict):
            raise ValueError("TXT/JSON dataset items must be objects")
        record = {str(key): _normalize_cell(value) for key, value in item.items()}
        records.append(record)
        for key in record.keys():
            if key not in seen_fields:
                seen_fields.add(key)
                fieldnames.append(key)
    if not fieldnames:
        raise ValueError("TXT/JSON dataset is missing object fields")
    return records, fieldnames


def _load_records_from_upload(filename: str, raw_bytes: bytes) -> tuple[list[dict[str, str]], list[str], str]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        records, fieldnames = _load_csv_records(raw_bytes)
        return records, fieldnames, "csv"
    if suffix == ".xlsx":
        records, fieldnames = _load_xlsx_records(raw_bytes)
        return records, fieldnames, "xlsx"
    if suffix in {".txt", ".json"}:
        records, fieldnames = _load_json_records(raw_bytes)
        return records, fieldnames, "json"
    raise ValueError("Only CSV, XLSX, TXT, and JSON datasets are supported")


def _ensure_field_exists(fieldnames: list[str], field_name: str, label: str) -> None:
    if field_name not in fieldnames:
        raise ValueError(f"Dataset missing {label}: {field_name}")


def _parse_expected_label_columns(value: list[str] | None) -> list[str]:
    if not value:
        return []
    parsed = [_normalize_cell(item) for item in value if _normalize_cell(item)]
    deduplicated: list[str] = []
    seen: set[str] = set()
    for item in parsed:
        lowered = item.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        deduplicated.append(item)
    return deduplicated


def normalize_compare_dataset(
    *,
    filename: str,
    raw_bytes: bytes,
    label_mode: LabelMode,
    text_column: str,
    expected_label_column: str | None = None,
    expected_label_columns: list[str] | None = None,
    positive_label_value: str | None = "1",
    sample_id_column: str | None = None,
    dataset_template: str | None = None,
) -> NormalizedDataset:
    records, fieldnames, source_format = _load_records_from_upload(filename, raw_bytes)

    text_column = _normalize_cell(text_column)
    if not text_column:
        raise ValueError("text_column is required")
    _ensure_field_exists(fieldnames, text_column, "text column")

    sample_id_column = _normalize_cell(sample_id_column) or None
    if sample_id_column:
        _ensure_field_exists(fieldnames, sample_id_column, "sample id column")

    normalized_rows: list[NormalizedCompareRow] = []
    rows_skipped = 0
    labels: list[str] = []

    if label_mode == "single_label":
        expected_label_column = _normalize_cell(expected_label_column)
        if not expected_label_column:
            raise ValueError("expected_label_column is required for single_label datasets")
        _ensure_field_exists(fieldnames, expected_label_column, "expected label column")
        task_type: Literal["single_label", "multi_label"] = "single_label"
        expected_label_columns = []
        positive_label_value = None
    elif label_mode == "multi_binary":
        expected_label_column = None
        expected_label_columns = _parse_expected_label_columns(expected_label_columns)
        if not expected_label_columns:
            raise ValueError("expected_label_columns must not be empty for multi_binary datasets")
        missing_columns = [column for column in expected_label_columns if column not in fieldnames]
        if missing_columns:
            raise ValueError(f"Dataset missing expected label columns: {', '.join(missing_columns)}")
        positive_label_value = _normalize_cell(positive_label_value) or "1"
        task_type = "multi_label"
    else:
        raise ValueError(f"Unsupported label_mode: {label_mode}")

    label_set: set[str] = set()
    positive_tokens = {
        token.lower()
        for token in (positive_label_value or "1").split(",")
        if token.strip()
    }

    for index, record in enumerate(records, start=1):
        text = _normalize_cell(record.get(text_column))
        if not text:
            rows_skipped += 1
            continue

        sample_id = _normalize_cell(record.get(sample_id_column)) if sample_id_column else None
        ground_truth_labels: list[str]
        if task_type == "single_label":
            raw_label = _normalize_label(record.get(expected_label_column or ""))
            ground_truth_labels = [raw_label] if raw_label else []
        else:
            ground_truth_labels = [
                column
                for column in expected_label_columns
                if _normalize_cell(record.get(column)).lower() in positive_tokens
            ]

        label_set.update(ground_truth_labels)
        normalized_rows.append(
            NormalizedCompareRow(
                row_index=index,
                sample_id=sample_id or None,
                text=text,
                ground_truth_labels=ground_truth_labels,
            )
        )

    if task_type == "single_label":
        labels = sorted(label_set)
        label_count = len(labels)
    else:
        labels = list(expected_label_columns or [])
        label_count = len(labels)

    return NormalizedDataset(
        dataset_name=filename,
        source_format=source_format,
        task_type=task_type,
        dataset_template=dataset_template,
        sample_id_column=sample_id_column,
        text_column=text_column,
        expected_label_column=expected_label_column,
        expected_label_columns=expected_label_columns or [],
        positive_label_value=positive_label_value,
        rows_processed=len(normalized_rows),
        rows_skipped=rows_skipped,
        label_count=label_count,
        labels=labels,
        rows=normalized_rows,
    )


def _normalized_prediction_tokens(
    *,
    emotion_code: str | None,
    emotion_name: str | None,
    raw_label: str | None,
) -> set[str]:
    tokens = {
        value.lower()
        for value in [emotion_code, emotion_name, raw_label]
        if value and value.strip()
    }
    return tokens


def _merge_label_mapping(label_mapping: dict[str, str | list[str]] | None) -> dict[str, list[str]]:
    merged = {key: list(values) for key, values in build_default_label_mapping().items()}
    normalized_custom = normalize_label_mapping(label_mapping)
    for key, values in normalized_custom.items():
        merged[key] = sorted(set(values + merged.get(key, [])))
    return merged


def _prediction_matches_labels(
    *,
    ground_truth_labels: list[str],
    emotion_code: str | None,
    emotion_name: str | None,
    raw_label: str | None,
    label_mapping: dict[str, list[str]],
) -> bool | None:
    normalized_truths = [label for label in ground_truth_labels if label.strip()]
    if not normalized_truths:
        return None
    projected_labels = _project_prediction_labels(
        dataset_labels=normalized_truths,
        emotion_code=emotion_code,
        emotion_name=emotion_name,
        raw_label=raw_label,
        label_mapping=label_mapping,
        include_fallback=False,
    )
    return bool(projected_labels)


def _project_prediction_label(
    *,
    dataset_labels: list[str],
    emotion_code: str | None,
    emotion_name: str | None,
    raw_label: str | None,
    label_mapping: dict[str, list[str]],
) -> str | None:
    prediction_tokens = _normalized_prediction_tokens(
        emotion_code=emotion_code,
        emotion_name=emotion_name,
        raw_label=raw_label,
    )
    if not prediction_tokens:
        return None

    for label in dataset_labels:
        normalized = label.lower()
        acceptable = {normalized}
        acceptable.update(label_mapping.get(normalized, []))
        if any(token in acceptable for token in prediction_tokens):
            return label
    return raw_label or emotion_name or emotion_code


def _project_prediction_labels(
    *,
    dataset_labels: list[str],
    emotion_code: str | None,
    emotion_name: str | None,
    raw_label: str | None,
    label_mapping: dict[str, list[str]],
    include_fallback: bool = True,
) -> list[str]:
    prediction_tokens = _normalized_prediction_tokens(
        emotion_code=emotion_code,
        emotion_name=emotion_name,
        raw_label=raw_label,
    )
    if not prediction_tokens:
        return []

    matched_labels: list[str] = []
    for label in dataset_labels:
        normalized = label.lower()
        acceptable = {normalized}
        acceptable.update(label_mapping.get(normalized, []))
        if any(token in acceptable for token in prediction_tokens):
            matched_labels.append(label)
    if matched_labels:
        return matched_labels

    if not include_fallback:
        return []

    fallback = raw_label or emotion_name or emotion_code
    return [fallback] if fallback else []


def _compute_single_label_metrics(
    *,
    comparison_rows: list[EmotionCompareComparisonRow],
    dataset_labels: list[str],
    label_mapping: dict[str, list[str]],
    side: Literal["baseline", "system"],
) -> EmotionCompareSingleMetrics:
    truth_labels: list[str] = []
    predicted_labels: list[str] = []
    correct = 0

    for row in comparison_rows:
        if len(row.groundTruthLabels) != 1:
            continue
        truth = row.groundTruthLabels[0]
        prediction = row.baselinePrediction if side == "baseline" else row.systemPrediction
        projected = _project_prediction_label(
            dataset_labels=dataset_labels,
            emotion_code=prediction.emotionCode,
            emotion_name=prediction.emotionName,
            raw_label=prediction.rawLabel,
            label_mapping=label_mapping,
        )
        truth_labels.append(truth)
        predicted_labels.append(projected or "__none__")
        if projected == truth:
            correct += 1

    support = len(truth_labels)
    if support == 0:
        return EmotionCompareSingleMetrics()

    label_space = sorted(set(dataset_labels) | set(truth_labels) | set(predicted_labels))
    per_label_f1: list[float] = []
    weighted_total = 0.0
    for label in label_space:
        tp = sum(1 for truth, pred in zip(truth_labels, predicted_labels) if truth == label and pred == label)
        fp = sum(1 for truth, pred in zip(truth_labels, predicted_labels) if truth != label and pred == label)
        fn = sum(1 for truth, pred in zip(truth_labels, predicted_labels) if truth == label and pred != label)
        precision = tp / (tp + fp) if tp + fp > 0 else 0.0
        recall = tp / (tp + fn) if tp + fn > 0 else 0.0
        f1 = (2 * precision * recall / (precision + recall)) if precision + recall > 0 else 0.0
        support_count = sum(1 for truth in truth_labels if truth == label)
        per_label_f1.append(f1)
        weighted_total += f1 * support_count

    return EmotionCompareSingleMetrics(
        accuracy=correct / support,
        macroF1=sum(per_label_f1) / len(per_label_f1) if per_label_f1 else 0.0,
        weightedF1=weighted_total / support if support else 0.0,
    )


def _compute_multi_label_metrics(
    *,
    comparison_rows: list[EmotionCompareComparisonRow],
    dataset_labels: list[str],
    label_mapping: dict[str, list[str]],
    side: Literal["baseline", "system"],
) -> EmotionCompareMultiMetrics:
    support = len(comparison_rows)
    if support == 0:
        return EmotionCompareMultiMetrics()

    exact_match_count = 0
    overlap_match_count = 0
    per_label_f1: list[float] = []
    label_metrics: list[EmotionComparePerLabelMetric] = []
    micro_tp = micro_fp = micro_fn = 0

    for row in comparison_rows:
        prediction = row.baselinePrediction if side == "baseline" else row.systemPrediction
        projected_set = set(
            _project_prediction_labels(
                dataset_labels=dataset_labels,
                emotion_code=prediction.emotionCode,
                emotion_name=prediction.emotionName,
                raw_label=prediction.rawLabel,
                label_mapping=label_mapping,
            )
        )
        truth_set = set(row.groundTruthLabels)
        if truth_set and projected_set == truth_set:
            exact_match_count += 1
        if truth_set and projected_set & truth_set:
            overlap_match_count += 1

    for label in dataset_labels:
        tp = fp = fn = 0
        support_count = 0
        for row in comparison_rows:
            prediction = row.baselinePrediction if side == "baseline" else row.systemPrediction
            projected_set = set(
                _project_prediction_labels(
                    dataset_labels=dataset_labels,
                    emotion_code=prediction.emotionCode,
                    emotion_name=prediction.emotionName,
                    raw_label=prediction.rawLabel,
                    label_mapping=label_mapping,
                )
            )
            truth_set = set(row.groundTruthLabels)
            truth_has = label in truth_set
            pred_has = label in projected_set
            if truth_has:
                support_count += 1
            if truth_has and pred_has:
                tp += 1
            elif not truth_has and pred_has:
                fp += 1
            elif truth_has and not pred_has:
                fn += 1

        micro_tp += tp
        micro_fp += fp
        micro_fn += fn
        precision = tp / (tp + fp) if tp + fp > 0 else 0.0
        recall = tp / (tp + fn) if tp + fn > 0 else 0.0
        f1 = (2 * precision * recall / (precision + recall)) if precision + recall > 0 else 0.0
        per_label_f1.append(f1)
        label_metrics.append(
            EmotionComparePerLabelMetric(
                label=label,
                precision=precision,
                recall=recall,
                f1=f1,
                support=support_count,
            )
        )

    micro_precision = micro_tp / (micro_tp + micro_fp) if micro_tp + micro_fp > 0 else 0.0
    micro_recall = micro_tp / (micro_tp + micro_fn) if micro_tp + micro_fn > 0 else 0.0
    micro_f1 = (
        2 * micro_precision * micro_recall / (micro_precision + micro_recall)
        if micro_precision + micro_recall > 0
        else 0.0
    )

    return EmotionCompareMultiMetrics(
        exactMatch=exact_match_count / support,
        overlapMatch=overlap_match_count / support,
        microF1=micro_f1,
        macroF1=sum(per_label_f1) / len(per_label_f1) if per_label_f1 else 0.0,
        labelWiseMetrics=label_metrics,
    )


def _build_predicted_labels(
    *,
    dataset_labels: list[str],
    prediction: EmotionComparePrediction,
    label_mapping: dict[str, list[str]],
) -> list[str]:
    return _project_prediction_labels(
            dataset_labels=dataset_labels,
            emotion_code=prediction.emotionCode,
            emotion_name=prediction.emotionName,
            raw_label=prediction.rawLabel,
            label_mapping=label_mapping,
        )


def _compute_binary_metrics(tp: int, fp: int, fn: int) -> tuple[float, float, float]:
    precision = tp / (tp + fp) if tp + fp > 0 else 0.0
    recall = tp / (tp + fn) if tp + fn > 0 else 0.0
    f1 = (2 * precision * recall / (precision + recall)) if precision + recall > 0 else 0.0
    return precision, recall, f1


def _compute_per_class_metrics_v2(
    *,
    comparison_rows: list[EmotionCompareComparisonRow],
    dataset_labels: list[str],
    task_type: Literal["single_label", "multi_label"],
) -> list[EmotionComparePerClassMetricItem]:
    metrics: list[EmotionComparePerClassMetricItem] = []
    for label in dataset_labels:
        baseline_tp = baseline_fp = baseline_fn = 0
        system_tp = system_fp = system_fn = 0
        support = 0

        for row in comparison_rows:
            truth_has = label in row.groundTruthLabels
            if truth_has:
                support += 1
            if task_type == "single_label":
                baseline_has = row.baselinePrediction.predictedLabels[:1] == [label]
                system_has = row.systemPrediction.predictedLabels[:1] == [label]
            else:
                baseline_has = label in row.baselinePrediction.predictedLabels
                system_has = label in row.systemPrediction.predictedLabels

            if truth_has and baseline_has:
                baseline_tp += 1
            elif baseline_has and not truth_has:
                baseline_fp += 1
            elif truth_has and not baseline_has:
                baseline_fn += 1

            if truth_has and system_has:
                system_tp += 1
            elif system_has and not truth_has:
                system_fp += 1
            elif truth_has and not system_has:
                system_fn += 1

        baseline_precision, baseline_recall, baseline_f1 = _compute_binary_metrics(
            baseline_tp, baseline_fp, baseline_fn
        )
        system_precision, system_recall, system_f1 = _compute_binary_metrics(
            system_tp, system_fp, system_fn
        )
        metrics.append(
            EmotionComparePerClassMetricItem(
                label=label,
                support=support,
                baselinePrecision=baseline_precision,
                baselineRecall=baseline_recall,
                baselineF1=baseline_f1,
                systemPrecision=system_precision,
                systemRecall=system_recall,
                systemF1=system_f1,
            )
        )
    return metrics


def _compute_confusion_analysis(
    *,
    comparison_rows: list[EmotionCompareComparisonRow],
    labels: list[str],
    task_type: Literal["single_label", "multi_label"],
) -> EmotionCompareConfusionAnalysis:
    if task_type != "single_label" or not labels:
        return EmotionCompareConfusionAnalysis(taskType=task_type, labels=labels, available=False)

    index_map = {label: index for index, label in enumerate(labels)}
    baseline_matrix = [[0 for _ in labels] for _ in labels]
    system_matrix = [[0 for _ in labels] for _ in labels]

    for row in comparison_rows:
        if len(row.groundTruthLabels) != 1:
            continue
        truth = row.groundTruthLabels[0]
        truth_index = index_map.get(truth)
        if truth_index is None:
            continue
        baseline_label = row.baselinePrediction.predictedLabels[0] if row.baselinePrediction.predictedLabels else None
        system_label = row.systemPrediction.predictedLabels[0] if row.systemPrediction.predictedLabels else None
        if baseline_label in index_map:
            baseline_matrix[truth_index][index_map[baseline_label]] += 1
        if system_label in index_map:
            system_matrix[truth_index][index_map[system_label]] += 1

    return EmotionCompareConfusionAnalysis(
        taskType=task_type,
        labels=labels,
        baselineMatrix=baseline_matrix,
        systemMatrix=system_matrix,
        available=True,
    )


def _compute_agreement_metrics(
    *,
    comparison_rows: list[EmotionCompareComparisonRow],
    labels: list[str],
    task_type: Literal["single_label", "multi_label"],
) -> EmotionCompareAgreementMetrics:
    if task_type != "single_label" or not labels:
        return EmotionCompareAgreementMetrics(taskType=task_type, available=False)

    support = 0
    baseline_pairs: list[tuple[str, str]] = []
    system_pairs: list[tuple[str, str]] = []
    for row in comparison_rows:
        if len(row.groundTruthLabels) != 1:
            continue
        truth = row.groundTruthLabels[0]
        baseline_pred = row.baselinePrediction.predictedLabels[0] if row.baselinePrediction.predictedLabels else "__none__"
        system_pred = row.systemPrediction.predictedLabels[0] if row.systemPrediction.predictedLabels else "__none__"
        baseline_pairs.append((truth, baseline_pred))
        system_pairs.append((truth, system_pred))
        support += 1

    def build_metric_set(pairs: list[tuple[str, str]]) -> EmotionCompareAgreementMetricSet:
        if support == 0:
            return EmotionCompareAgreementMetricSet()
        po = sum(1 for truth, pred in pairs if truth == pred) / support
        truth_counts = {label: 0 for label in labels}
        pred_counts = {label: 0 for label in labels}
        for truth, pred in pairs:
            if truth in truth_counts:
                truth_counts[truth] += 1
            if pred in pred_counts:
                pred_counts[pred] += 1
        pe = sum((truth_counts[label] / support) * (pred_counts[label] / support) for label in labels)
        kappa = (po - pe) / (1 - pe) if pe < 1 else 1.0
        return EmotionCompareAgreementMetricSet(
            overallAgreement=po,
            observedAgreement=po,
            expectedAgreement=pe,
            cohensKappa=kappa,
        )

    return EmotionCompareAgreementMetrics(
        taskType=task_type,
        baseline=build_metric_set(baseline_pairs),
        system=build_metric_set(system_pairs),
        available=True,
    )


def _compute_intensity_metrics(
    *,
    comparison_rows: list[EmotionCompareComparisonRow],
) -> EmotionCompareIntensityMetrics:
    del comparison_rows
    return EmotionCompareIntensityMetrics(
        baseline=EmotionCompareIntensityMetricSet(),
        system=EmotionCompareIntensityMetricSet(),
        available=False,
        note="当前公开数据集缺少可靠强度真值，第二阶段先保留结构与空值语义。",
    )


def _build_export_csv(comparison_rows: list[EmotionCompareComparisonRow]) -> str:
    csv_output = io.StringIO()
    writer = csv.DictWriter(
        csv_output,
        fieldnames=[
            "rowIndex",
            "sampleId",
            "text",
            "groundTruthLabels",
            "baselineEmotionCode",
            "baselineEmotionName",
            "baselinePredictedLabels",
            "baselineIntensity",
            "baselineConfidence",
            "systemEmotionCode",
            "systemEmotionName",
            "systemPredictedLabels",
            "systemIntensity",
            "systemConfidence",
            "baselineMatched",
            "systemMatched",
            "winner",
        ],
    )
    writer.writeheader()
    for row in comparison_rows:
        writer.writerow(
            {
                "rowIndex": row.rowIndex,
                "sampleId": row.sampleId or "",
                "text": row.text,
                "groundTruthLabels": "|".join(row.groundTruthLabels),
                "baselineEmotionCode": row.baselinePrediction.emotionCode or "",
                "baselineEmotionName": row.baselinePrediction.emotionName or "",
                "baselinePredictedLabels": "|".join(row.baselinePrediction.predictedLabels),
                "baselineIntensity": row.baselinePrediction.intensity or "",
                "baselineConfidence": row.baselinePrediction.confidence if row.baselinePrediction.confidence is not None else "",
                "systemEmotionCode": row.systemPrediction.emotionCode or "",
                "systemEmotionName": row.systemPrediction.emotionName or "",
                "systemPredictedLabels": "|".join(row.systemPrediction.predictedLabels),
                "systemIntensity": row.systemPrediction.intensity or "",
                "systemConfidence": row.systemPrediction.confidence if row.systemPrediction.confidence is not None else "",
                "baselineMatched": "" if row.baselineMatched is None else row.baselineMatched,
                "systemMatched": "" if row.systemMatched is None else row.systemMatched,
                "winner": row.winner,
            }
        )
    return csv_output.getvalue()


async def _analyze_baseline_row(
    row: NormalizedCompareRow,
    analyzer: TextAnalyzer,
) -> EmotionCompareBaselineRow:
    analysis = await analyzer.analyze(user_message=row.text)
    detail = analysis.emotionDetail
    return EmotionCompareBaselineRow(
        rowIndex=row.row_index,
        sampleId=row.sample_id,
        text=row.text,
        predictedEmotionCode=detail.code if detail else None,
        predictedEmotionName=detail.name if detail else None,
        predictedIntensity=detail.intensity if detail else None,
        confidence=detail.confidence if detail else None,
        rawLabel=analysis.emotion,
    )


async def _analyze_system_row(
    row: NormalizedCompareRow,
    analyzer: TextAnalyzer,
    profile_state: UserProfile,
) -> tuple[EmotionCompareSystemRow, UserProfile]:
    analysis = await analyzer.analyze(
        user_message=row.text,
        current_profile=profile_state,
    )
    detail = analysis.emotionDetail
    updated_profile = _apply_delta(
        profile_state,
        analysis.delta.cognition,
        analysis.delta.affect,
        analysis.delta.behavior,
    )
    system_row = EmotionCompareSystemRow(
        rowIndex=row.row_index,
        sampleId=row.sample_id,
        text=row.text,
        predictedEmotionCode=detail.code if detail else None,
        predictedEmotionName=detail.name if detail else None,
        predictedIntensity=detail.intensity if detail else None,
        confidence=detail.confidence if detail else None,
        profileBefore=EmotionCompareProfileSnapshot(
            cognition=profile_state.cognition,
            affect=profile_state.affect,
            behavior=profile_state.behavior,
        ),
        profileAfter=EmotionCompareProfileSnapshot(
            cognition=updated_profile.cognition,
            affect=updated_profile.affect,
            behavior=updated_profile.behavior,
        ),
        delta=EmotionCompareDelta(
            cognition=analysis.delta.cognition,
            affect=analysis.delta.affect,
            behavior=analysis.delta.behavior,
        ),
        contextUsed=EmotionCompareContextUsed(dialogue=False, profile=True, knowledge=False),
    )
    return system_row, updated_profile


def _prediction_from_baseline_row(row: EmotionCompareBaselineRow) -> EmotionComparePrediction:
    return EmotionComparePrediction(
        emotionCode=row.predictedEmotionCode,
        emotionName=row.predictedEmotionName,
        intensity=row.predictedIntensity,
        confidence=row.confidence,
        rawLabel=row.rawLabel,
    )


def _prediction_from_system_row(row: EmotionCompareSystemRow) -> EmotionComparePrediction:
    return EmotionComparePrediction(
        emotionCode=row.predictedEmotionCode,
        emotionName=row.predictedEmotionName,
        intensity=row.predictedIntensity,
        confidence=row.confidence,
        rawLabel=row.predictedEmotionName,
    )


async def analyze_compare_dataset(
    *,
    dataset: NormalizedDataset,
    preview_limit: int = 50,
    label_mapping: dict[str, str | list[str]] | None = None,
    provider: BaseProvider | None = None,
) -> EmotionCompareAnalyzeResponse:
    limited_rows = dataset.rows[:preview_limit]
    baseline_analyzer = TextAnalyzer(provider=provider)
    system_analyzer = TextAnalyzer(provider=provider)
    merged_label_mapping = _merge_label_mapping(label_mapping)

    baseline_rows: list[EmotionCompareBaselineRow] = []
    system_rows: list[EmotionCompareSystemRow] = []
    comparison_rows: list[EmotionCompareComparisonRow] = []

    profile_states: dict[str, UserProfile] = {}

    for row in limited_rows:
        profile_key = row.sample_id or "__default_profile__"
        profile_before = profile_states.get(profile_key, _default_profile())

        try:
            baseline_row = await _analyze_baseline_row(row, baseline_analyzer)
        except Exception as exc:
            logger.warning("Baseline analysis failed for row %s: %s", row.row_index, exc)
            baseline_row = EmotionCompareBaselineRow(
                rowIndex=row.row_index,
                sampleId=row.sample_id,
                text=row.text,
            )

        try:
            system_row, updated_profile = await _analyze_system_row(row, system_analyzer, profile_before)
            profile_states[profile_key] = updated_profile
        except Exception as exc:
            logger.warning("System analysis failed for row %s: %s", row.row_index, exc)
            system_row = EmotionCompareSystemRow(
                rowIndex=row.row_index,
                sampleId=row.sample_id,
                text=row.text,
                profileBefore=EmotionCompareProfileSnapshot(
                    cognition=profile_before.cognition,
                    affect=profile_before.affect,
                    behavior=profile_before.behavior,
                ),
                profileAfter=EmotionCompareProfileSnapshot(
                    cognition=profile_before.cognition,
                    affect=profile_before.affect,
                    behavior=profile_before.behavior,
                ),
                delta=EmotionCompareDelta(cognition=0, affect=0, behavior=0),
                contextUsed=EmotionCompareContextUsed(dialogue=False, profile=True, knowledge=False),
            )

        baseline_rows.append(baseline_row)
        system_rows.append(system_row)

        baseline_prediction = _prediction_from_baseline_row(baseline_row)
        system_prediction = _prediction_from_system_row(system_row)
        baseline_prediction.predictedLabels = _build_predicted_labels(
            dataset_labels=dataset.labels,
            prediction=baseline_prediction,
            label_mapping=merged_label_mapping,
        )
        system_prediction.predictedLabels = _build_predicted_labels(
            dataset_labels=dataset.labels,
            prediction=system_prediction,
            label_mapping=merged_label_mapping,
        )
        baseline_matched = _prediction_matches_labels(
            ground_truth_labels=row.ground_truth_labels,
            emotion_code=baseline_prediction.emotionCode,
            emotion_name=baseline_prediction.emotionName,
            raw_label=baseline_prediction.rawLabel,
            label_mapping=merged_label_mapping,
        )
        system_matched = _prediction_matches_labels(
            ground_truth_labels=row.ground_truth_labels,
            emotion_code=system_prediction.emotionCode,
            emotion_name=system_prediction.emotionName,
            raw_label=system_prediction.rawLabel,
            label_mapping=merged_label_mapping,
        )

        winner: Literal["baseline", "system", "tie", "none"] = "none"
        if baseline_matched is True and system_matched is not True:
            winner = "baseline"
        elif system_matched is True and baseline_matched is not True:
            winner = "system"
        elif system_matched is True and baseline_matched is True:
            winner = "tie"

        comparison_rows.append(
            EmotionCompareComparisonRow(
                rowIndex=row.row_index,
                sampleId=row.sample_id,
                text=row.text,
                groundTruthLabels=row.ground_truth_labels,
                baselinePrediction=baseline_prediction,
                systemPrediction=system_prediction,
                baselineMatched=baseline_matched,
                systemMatched=system_matched,
                winner=winner,
            )
        )

    if dataset.task_type == "single_label":
        summary_metrics = EmotionCompareSummaryMetrics(
            taskType="single_label",
            support=dataset.rows_processed,
            labelCount=dataset.label_count,
            labels=dataset.labels,
            baseline=_compute_single_label_metrics(
                comparison_rows=comparison_rows,
                dataset_labels=dataset.labels,
                label_mapping=merged_label_mapping,
                side="baseline",
            ),
            system=_compute_single_label_metrics(
                comparison_rows=comparison_rows,
                dataset_labels=dataset.labels,
                label_mapping=merged_label_mapping,
                side="system",
            ),
        )
    else:
        summary_metrics = EmotionCompareSummaryMetrics(
            taskType="multi_label",
            support=dataset.rows_processed,
            labelCount=dataset.label_count,
            labels=dataset.labels,
            baseline=_compute_multi_label_metrics(
                comparison_rows=comparison_rows,
                dataset_labels=dataset.labels,
                label_mapping=merged_label_mapping,
                side="baseline",
            ),
            system=_compute_multi_label_metrics(
                comparison_rows=comparison_rows,
                dataset_labels=dataset.labels,
                label_mapping=merged_label_mapping,
                side="system",
            ),
        )

    stem = Path(dataset.dataset_name).stem
    return EmotionCompareAnalyzeResponse(
        datasetInfo=EmotionCompareDatasetInfo(
            datasetName=dataset.dataset_name,
            sourceFormat=dataset.source_format,
            taskType=dataset.task_type,
            datasetTemplate=dataset.dataset_template,
            sampleIdColumn=dataset.sample_id_column,
            textColumn=dataset.text_column,
            expectedLabelColumn=dataset.expected_label_column,
            expectedLabelColumns=dataset.expected_label_columns,
            positiveLabelValue=dataset.positive_label_value,
            rowsProcessed=dataset.rows_processed,
            rowsSkipped=dataset.rows_skipped,
            labelCount=dataset.label_count,
            labels=dataset.labels,
        ),
        baselineRows=baseline_rows,
        systemRows=system_rows,
        comparisonRows=comparison_rows,
        summaryMetrics=summary_metrics,
        exportArtifacts=EmotionCompareExportArtifacts(
            comparisonCsvFileName=f"{stem}_compare_results.csv",
            comparisonCsvContent=_build_export_csv(comparison_rows),
            resultJsonFileName=f"{stem}_compare_results.json",
        ),
    )


async def analyze_compare_dataset_v2(
    *,
    dataset: NormalizedDataset,
    preview_limit: int = 50,
    label_mapping: dict[str, str | list[str]] | None = None,
    provider: BaseProvider | None = None,
) -> EmotionCompareAnalyzeResponseV2:
    base = await analyze_compare_dataset(
        dataset=dataset,
        preview_limit=preview_limit,
        label_mapping=label_mapping,
        provider=provider,
    )

    return EmotionCompareAnalyzeResponseV2(
        base=base,
        perClassMetrics=_compute_per_class_metrics_v2(
            comparison_rows=base.comparisonRows,
            dataset_labels=base.datasetInfo.labels,
            task_type=base.datasetInfo.taskType,
        ),
        confusionAnalysis=_compute_confusion_analysis(
            comparison_rows=base.comparisonRows,
            labels=base.datasetInfo.labels,
            task_type=base.datasetInfo.taskType,
        ),
        agreementMetrics=_compute_agreement_metrics(
            comparison_rows=base.comparisonRows,
            labels=base.datasetInfo.labels,
            task_type=base.datasetInfo.taskType,
        ),
        intensityMetrics=_compute_intensity_metrics(
            comparison_rows=base.comparisonRows,
        ),
    )


def build_emotion_compare_skeleton_response(
    *,
    dataset: NormalizedDataset,
    preview_limit: int = 50,
) -> EmotionCompareAnalyzeResponse:
    limited_rows = dataset.rows[:preview_limit]

    baseline_rows = [
        EmotionCompareBaselineRow(
            rowIndex=row.row_index,
            sampleId=row.sample_id,
            text=row.text,
        )
        for row in limited_rows
    ]
    system_rows = [
        EmotionCompareSystemRow(
            rowIndex=row.row_index,
            sampleId=row.sample_id,
            text=row.text,
            profileBefore=EmotionCompareProfileSnapshot(cognition=50, affect=50, behavior=50),
            profileAfter=EmotionCompareProfileSnapshot(cognition=50, affect=50, behavior=50),
        )
        for row in limited_rows
    ]
    comparison_rows = [
        EmotionCompareComparisonRow(
            rowIndex=row.row_index,
            sampleId=row.sample_id,
            text=row.text,
            groundTruthLabels=row.ground_truth_labels,
        )
        for row in limited_rows
    ]

    if dataset.task_type == "single_label":
        summary_metrics = EmotionCompareSummaryMetrics(
            taskType="single_label",
            support=dataset.rows_processed,
            labelCount=dataset.label_count,
            labels=dataset.labels,
            baseline=EmotionCompareSingleMetrics(),
            system=EmotionCompareSingleMetrics(),
        )
    else:
        summary_metrics = EmotionCompareSummaryMetrics(
            taskType="multi_label",
            support=dataset.rows_processed,
            labelCount=dataset.label_count,
            labels=dataset.labels,
            baseline=EmotionCompareMultiMetrics(),
            system=EmotionCompareMultiMetrics(),
        )

    csv_output = io.StringIO()
    writer = csv.DictWriter(
        csv_output,
        fieldnames=[
            "rowIndex",
            "sampleId",
            "text",
            "groundTruthLabels",
            "baselineEmotionName",
            "systemEmotionName",
            "baselineMatched",
            "systemMatched",
            "winner",
        ],
    )
    writer.writeheader()
    for row in comparison_rows:
        writer.writerow(
            {
                "rowIndex": row.rowIndex,
                "sampleId": row.sampleId or "",
                "text": row.text,
                "groundTruthLabels": "|".join(row.groundTruthLabels),
                "baselineEmotionName": row.baselinePrediction.emotionName or "",
                "systemEmotionName": row.systemPrediction.emotionName or "",
                "baselineMatched": "" if row.baselineMatched is None else row.baselineMatched,
                "systemMatched": "" if row.systemMatched is None else row.systemMatched,
                "winner": row.winner,
            }
        )

    stem = Path(dataset.dataset_name).stem
    return EmotionCompareAnalyzeResponse(
        datasetInfo=EmotionCompareDatasetInfo(
            datasetName=dataset.dataset_name,
            sourceFormat=dataset.source_format,
            taskType=dataset.task_type,
            datasetTemplate=dataset.dataset_template,
            sampleIdColumn=dataset.sample_id_column,
            textColumn=dataset.text_column,
            expectedLabelColumn=dataset.expected_label_column,
            expectedLabelColumns=dataset.expected_label_columns,
            positiveLabelValue=dataset.positive_label_value,
            rowsProcessed=dataset.rows_processed,
            rowsSkipped=dataset.rows_skipped,
            labelCount=dataset.label_count,
            labels=dataset.labels,
        ),
        baselineRows=baseline_rows,
        systemRows=system_rows,
        comparisonRows=comparison_rows,
        summaryMetrics=summary_metrics,
        exportArtifacts=EmotionCompareExportArtifacts(
            comparisonCsvFileName=f"{stem}_compare_results.csv",
            comparisonCsvContent=csv_output.getvalue(),
            resultJsonFileName=f"{stem}_compare_results.json",
        ),
    )
